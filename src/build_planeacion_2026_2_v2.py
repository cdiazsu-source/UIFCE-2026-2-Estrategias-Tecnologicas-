# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="7A1F2B", end_color="7A1F2B", fill_type="solid")
CAT_FILL = PatternFill(start_color="C9A0A6", end_color="C9A0A6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="7A1F2B", size=14)
SUB_FONT = Font(italic=True, color="595959", size=10)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER_TOP = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill(start_color="F7F0F1", end_color="F7F0F1", fill_type="solid")
PRIORITY_FILL = PatternFill(start_color="FCE9CE", end_color="FCE9CE", fill_type="solid")

# ---------------------------------------------------------------------------
# Contenido maestro: una entrada por idea, reutilizada en ambas hojas
# ---------------------------------------------------------------------------
IDEAS = [
    dict(
        id="redes-instagram",
        cat="Redes y canales — CRÍTICO",
        idea="Recuperación o recreación de la cuenta de Instagram @uifce_un",
        hacer="Agotar primero las vías de recuperación de la cuenta @uifce_un (verificación de identidad ante Meta, apoyo de Imagen Institucional/Unimedios, revisión de correos y dispositivos con sesión activa) e insistir por distintos canales con la persona con la que se ha intentado retomar contacto. Si no se logra recuperar en un plazo máximo de 2-3 semanas, escalar pidiendo la intermediación directa del profesor a cargo del área y, en paralelo, crear una cuenta nueva oficial, tramitar su oficialización ante la Oficina de Medios Digitales de la UNAL (mismo procedimiento ya usado con Instagram y LinkedIn en 2025-2) y migrar de inmediato los contenidos disponibles.",
        espera="Contar, a más tardar en la semana 4 del semestre, con una cuenta de Instagram activa y oficial (recuperada o nueva) desde la cual retomar el calendario editorial sin perder la continuidad de la estrategia de difusión.",
        fundamento="Instagram es, según el diagnóstico 2025-2 → 2026-1, el canal principal de la UIFCE: concentró 25 reels con 83.649 visualizaciones y 2.323 interacciones en el último semestre. Su pérdida es una contingencia reportada por el área que pone en riesgo directo la ejecución de todo el portafolio de comunicación 2026-2. El video de empalme con la líder saliente confirma que persiste incertidumbre real sobre qué ocurrió con la cuenta y sobre la identidad de la persona contactada, por lo que no conviene depender de una sola vía de gestión.",
        entregables="* Reporte de gestión de recuperación ante Meta.\n* Cuenta activa (recuperada o nueva) y, si aplica, trámite de oficialización iniciado.\n* Migración de contenidos disponibles.",
        priority=True,
    ),
    dict(
        id="redes-linkedin",
        cat="Redes y canales — PRIORITARIO",
        idea="Posicionamiento prioritario de LinkedIn UIFCE",
        hacer="Definir LinkedIn como canal prioritario del semestre: construir un calendario editorial propio (logros de monitores y egresados, proyectos de estudio, contenido profesional/técnico), aumentar la frecuencia de publicación frente a 2026-1 y evaluar la migración de contenido desde la cuenta personal hacia la cuenta empresa de la Unidad, pendiente según la Guía de Empalme.",
        espera="Superar de manera sostenida las cifras de 2026-1 (7 publicaciones, 1.491 impresiones, 42 reacciones), consolidando a LinkedIn como el canal de posicionamiento profesional e institucional de la UIFCE ante egresados, empresas y otras dependencias académicas.",
        fundamento="LinkedIn ya cuenta con oficialización institucional obtenida en 2026-1 y mostró el mayor crecimiento relativo del semestre (de una red subutilizada a un canal con métricas propias). Ante la contingencia de Instagram, se prioriza como canal estable adicional para no concentrar el riesgo de difusión en un solo canal.",
        entregables="* Calendario editorial LinkedIn.\n* Publicaciones mensuales mínimas.\n* Decisión sobre migración a cuenta empresa.",
        priority=True,
    ),
    dict(
        id="redes-tiktok",
        cat="Redes y canales — NUEVO",
        idea="Creación de la cuenta de TikTok UIFCE",
        hacer="Solicitar ante Imagen Institucional/Unimedios los lineamientos para cuentas institucionales en nuevas plataformas, crear y oficializar la cuenta de TikTok siguiendo el mismo trámite usado en 2026-1 para Instagram y LinkedIn, y definir un formato de contenido propio (no una simple réplica de los reels de Instagram). Complementar el lanzamiento con difusión \"análoga\" presencial (volantes/folletos entregados directamente a estudiantes en la Facultad, siguiendo la lógica de negocios cercanos que reparten volantes con promociones), en lugar de depender solo de que la audiencia encuentre la cuenta por su cuenta en redes.",
        espera="Contar con una cuenta de TikTok oficial y activa hacia la mitad del semestre, con un primer lote de publicaciones que permita evaluar su aporte real de alcance antes de comprometer más capacidad de producción del área.",
        fundamento="Los hacks informáticos y los reels demostraron en 2026-1 que el formato de video corto es el de mayor alcance de la UIFCE (23.029 visualizaciones en hacks, 83.649 en reels). TikTok es la plataforma nativa de ese formato y amplía el alcance hacia la población estudiantil que la usa como canal principal de consumo de contenido.",
        entregables="* Cuenta creada y oficializada.\n* Lineamiento de formato propio para TikTok.\n* Primer lote de publicaciones.",
        priority=False,
    ),
    dict(
        id="redes-youtube",
        cat="Redes y canales",
        idea="Consolidación y cierre de oficialización de YouTube UIFCE",
        hacer="Cerrar el trámite de oficialización de YouTube ante la Oficina de Medios Digitales UNAL, en curso desde 2025-2, resolviendo los inconvenientes de titularidad de cuenta ya identificados. Definir un calendario editorial propio (tutoriales, hacks en formato largo, videos institucionales) y coordinar con Virtualización la incrustación de videos de cursos.",
        espera="Trámite de oficialización cerrado y canal activo con publicaciones propias del área durante el semestre, además de servir como repositorio de los videos de cursos virtualizados que produce Virtualización.",
        fundamento="El informe de gestión 2025-2 dejó este trámite en curso; sin cerrarlo, el canal no cuenta con respaldo institucional formal pese a ya tener suscriptores y contenido de Virtualización dependiendo de él.",
        entregables="* Trámite de oficialización cerrado.\n* Calendario editorial YouTube.\n* Videos de Virtualización incrustados en Moodle.",
        priority=False,
    ),
    dict(
        id="normativa-terminos-condiciones",
        cat="Normativa",
        idea="Cierre de Términos y Condiciones de Estrategias Tecnológicas",
        hacer="Redactar el documento formal de Términos y Condiciones de ET (microtalleres, microeventos, Hackatón y otros eventos), definiendo con claridad las penalizaciones por inasistencia y dando continuidad al Manual de Microtalleres y Microeventos ya vigente. Presentarlo a Coordinación para aval.",
        espera="Documento aprobado y difundido a la comunidad UIFCE antes del cierre del primer tercio del semestre, de forma que rija desde los primeros microtalleres de 2026-2.",
        fundamento="La planeación 2026-1 dejó este ítem en estado pendiente, y el informe de gestión 2025-2 identificó explícitamente la falta de claridad en penalizaciones por inasistencia como una dificultad operativa del área.",
        entregables="* Documento de TyC aprobado.\n* Reglamentación de penalizaciones.\n* Difusión del documento.",
        priority=False,
    ),
    dict(
        id="produccion-piezas-audiovisual",
        cat="Producción de contenido",
        idea="Piezas gráficas y material audiovisual",
        hacer="Mantener la producción semanal de piezas y contenido audiovisual bajo un calendario editorial único (referencia histórica: un reel por semana y una pieza cada 15 días), priorizando calidad sobre volumen, y verificar la vigencia de la licencia de Adobe ya instalada en el equipo de la unidad usado por el monitor de artes (evitar depender de licencias personales). Retomar y terminar el material institucional pendiente identificado en el empalme: video de bienvenida tipo \"trailer\" para nuevos integrantes, ampliación del video de testimonios de ex-monitores (solo 3 grabados y sin editar), edición de los dos videos de apoyo académico ya grabados, y uso de los bloopers guardados para un evento de cierre de semestre.",
        espera="Un flujo constante de piezas aprobadas por el grupo de \"Piezas Redes Sociales\" (visto bueno de los coordinadores), sin los cuellos de botella de licencias reportados en 2025-2, y al menos dos de las piezas institucionales pendientes (trailer, ex-monitores, apoyo académico) terminadas durante el semestre.",
        fundamento="El diagnóstico identificó como lección aprendida que la cantidad de piezas no garantiza impacto, y como brecha la ausencia de licencias institucionales de software de diseño y edición. El video de empalme detalla piezas institucionales de alto potencial (especialmente el video de ex-monitores, señalado como palanca de atracción de nuevos monitores) que quedaron sin terminar en 2025-2.",
        entregables="* Calendario editorial 2026-2.\n* Piezas y videos publicados semana a semana.\n* Verificación de licencias institucionales (Adobe).\n* Al menos 2 piezas institucionales pendientes terminadas (trailer / ex-monitores / apoyo académico).",
        priority=False,
    ),
    dict(
        id="produccion-hacks-informaticos",
        cat="Producción de contenido",
        idea="Repositorio de Hacks Informáticos",
        hacer="Continuar la producción y publicación de hacks informáticos aplicando la rúbrica de calificación ya creada en 2026-1 (duración estándar, vigencia, calidad audiovisual). Depurar contenidos obsoletos y mantener actualizado el repositorio interno.",
        espera="Mantener o superar el alcance logrado en 2026-1 (23.029 visualizaciones y 526 \"me gusta\" en 11 hacks), con un repositorio depurado y vigente.",
        fundamento="Los hacks informáticos son, junto con los reels, el formato de mayor tracción medido en el diagnóstico 2025-2 → 2026-1.",
        entregables="* Nuevos hacks publicados.\n* Repositorio interno depurado y actualizado.",
        priority=False,
    ),
    dict(
        id="eventos-microtalleres",
        cat="Eventos",
        idea="Microtalleres UIFCE",
        hacer="Ejecutar un máximo de 5 microtalleres en el semestre (idealmente 3), priorizando calidad y asistencia efectiva sobre cantidad, y evitando programarlos en las últimas semanas del calendario académico. Implementar certificados de participación desde el primer microtaller (misma plantilla usada para Cursos Libres), confirmar con Juan Martínez (ex-monitor de geología) si dicta el microtaller ya diseñado de \"Análisis espacial aplicado a las ciencias económicas\" (5 sesiones), y confirmar con Jonny el horario del microtaller de Introducción a la Lógica de Programación.",
        espera="Mayor asistencia efectiva por microtaller, con un tope de 5 microtalleres en el semestre frente a los 10 de 2025-2, certificados de participación entregados desde el primer microtaller dictado, y al menos un piloto en modalidad virtual/híbrida.",
        fundamento="Lección aprendida directa del diagnóstico: alta inscripción con baja asistencia efectiva, y menor participación en los talleres realizados al cierre del calendario académico tras el paro. El video de empalme confirma este patrón (hasta 2% de asistencia en los últimos microtalleres de 2025-2) y añade la certificación de participación como demanda recurrente de los asistentes no resuelta hasta ahora.",
        entregables="* Cronograma de microtalleres (máx. 5).\n* Registro de asistencia efectiva.\n* Certificados de participación entregados.\n* Piloto de modalidad virtual/híbrida.",
        priority=False,
    ),
    dict(
        id="continuidad-blog",
        cat="Proyectos en continuidad",
        idea="Blog UIFCE",
        hacer="Aprobar y ejecutar la propuesta de difusión del blog, pendiente desde 2026-1, coordinando un cronograma conjunto con Gestión del Conocimiento (GC): GC cura y redacta los resúmenes de proyectos de estudio, ET diseña la plataforma y acompaña el despliegue.",
        espera="Blog con contenido y difusión activos durante el semestre, con la delimitación de funciones frente a GC funcionando en la práctica.",
        fundamento="La estrategia de difusión del blog quedó pendiente desde 2026-1; la delimitación de responsabilidades frente a GC ya se fijó en el Manual de Funciones y Estandarización.",
        entregables="* Estrategia de difusión aprobada y ejecutada.\n* Cronograma conjunto ET-GC.",
        priority=False,
    ),
    dict(
        id="continuidad-micrositio",
        cat="Proyectos en continuidad",
        idea="Micrositio UIFCE",
        hacer="Actualizar fotos, biografías y estructura de contenido del equipo vigente (coordinando con Comunicaciones una sesión de fotos institucionales, individuales y grupales por área), gestionando los consentimientos de imagen mediante el script ya existente en el Drive de ET (genera automáticamente un documento de consentimiento por persona a partir de la lista de monitores). Evaluar con Desarrollo (Johnny) la viabilidad técnica de que la foto de cada integrante del equipo enlace a su perfil de LinkedIn al hacer clic, como mecanismo adicional de posicionamiento del canal.",
        espera="Micrositio con información del equipo vigente para el semestre, desplegado sin retrasos técnicos por falta de coordinación con DS, y una respuesta clara de Desarrollo sobre la viabilidad del enlace a LinkedIn en las fotos del equipo.",
        fundamento="La Guía de Empalme señala este proceso como compartido: ET aporta el contenido y DS ejecuta el desarrollo y despliegue técnico. El video de empalme documenta que esta idea de vincular fotos a LinkedIn ya tuvo una prueba informal con buena recepción de Desarrollo, y que el proceso de consentimientos ya cuenta con un script funcional que solo requiere mantenerse actualizado.",
        entregables="* Fotos y biografías actualizadas.\n* Consentimientos de imagen gestionados vía script.\n* Micrositio desplegado.\n* Definición sobre el enlace foto-LinkedIn con Desarrollo.",
        priority=False,
    ),
    dict(
        id="produccion-senaletica-salas",
        cat="Producción de contenido",
        idea="Vigilancia y piezas de software disponible en cada sala",
        hacer="Mantener actualizadas las piezas gráficas de disponibilidad de software y salas en los televisores de oficina y pasillo, evitando el lenguaje restrictivo (tipo \"prohibido\") señalado en la reunión final 2026-1.",
        espera="Señalética de salas vigente durante todo el semestre y mejor recibida por los usuarios.",
        fundamento="Comentario explícito de la reunión final de 2026-1 sobre el tono restrictivo de las piezas anteriores.",
        entregables="* Piezas actualizadas por sala.\n* Repositorio de piezas en Drive.",
        priority=False,
    ),
    dict(
        id="cursoslibres-linea-grafica",
        cat="Acompañamiento a Cursos Libres",
        idea="Línea gráfica de Cursos Libres",
        hacer="Completar el despliegue de la nueva línea gráfica modular en medios físicos y digitales, y coordinar con Cursos Libres (CL) el mantenimiento de la cartelera y el diseño de plantillas de certificados conforme a Imagen Institucional.",
        espera="Línea gráfica desplegada al 100% en todos los medios, y cartelera de CL corregida y vigente.",
        fundamento="El despliegue de la nueva línea gráfica quedó en curso al cierre de 2026-1, a la espera de confirmación de logo.",
        entregables="* Línea gráfica desplegada.\n* Cartelera de CL corregida.\n* Plantilla de certificados aprobada.",
        priority=False,
    ),
    dict(
        id="estrategicos-extension-solidaria",
        cat="Proyectos estratégicos",
        idea="Extensión Solidaria",
        hacer="Presentar la propuesta técnica y presupuestal completa (metodología, población objetivo identificada a través de Bienestar Universitario, cotizaciones ya cerradas de kit escolar con Doneta y de refrigerio con Colombia Gourmet) ante Vicedecanatura para obtener el aval institucional. Esperar a que se complete el cambio de Vicedecanatura ya en curso antes de formalizar el trámite presupuestal, para no arriesgar compromisos con la administración saliente; explorar en paralelo la vía de una convocatoria de extensión asociada a problemáticas de Bogotá con apoyo de la Vicedecanatura de Investigación y Extensión (contacto: Sandra Carlos Vargas). Una vez aprobada, iniciar la gestión de ejecución.",
        espera="Aval institucional obtenido y, de ser posible dentro del semestre, ejecución de la capacitación en Excel dirigida a madres cabeza de familia (40 participantes, 15 horas).",
        fundamento="La propuesta técnica y las cotizaciones ya están completas desde 2026-1; solo falta el aval de Vicedecanatura para avanzar a ejecución. El video de empalme confirma el aval técnico del profesor Henry y precisa que la recomendación explícita es esperar la transición de Vicedecanatura antes de tramitar el presupuesto.",
        entregables="* Propuesta formal presentada.\n* Aval de Vicedecanatura.\n* Cotizaciones cerradas.",
        priority=False,
    ),
    dict(
        id="produccion-carteleras",
        cat="Producción de contenido",
        idea="Carteleras UIFCE (física y digital)",
        hacer="Mantener la proyección de piezas en las carteleras digitales de los edificios 310 y 311 en articulación con la Dependencia de Comunicaciones de la Facultad, y conservar y actualizar la cartelera física renovada en 2025-2.",
        espera="Presencia física constante y renovada de la UIFCE en los espacios de mayor circulación de la Facultad.",
        fundamento="La articulación con Comunicaciones FCE ya se logró en 2025-2 y debe mantenerse activa como canal complementario a las redes digitales.",
        entregables="* Piezas proyectadas en carteleras digitales 310/311.\n* Cartelera física actualizada.",
        priority=False,
    ),
    dict(
        id="eventos-semana-uifce",
        cat="Eventos / Proyectos estratégicos",
        idea="Semana UIFCE — primera edición",
        hacer="Ejecutar la primera edición del evento (septiembre), articulada con la Semana de Investigación de la Vicedecanatura de Investigación y Extensión (VIE, contacto Sandra Carlos Vargas), con tres componentes: (1) Hackatón en Bizagi a tres días no consecutivos (capacitación, entrega del caso, y sustentación una semana después), en equipos de 4 y con profesores como jurados, coordinada con Gau y Ángela, quienes ya tienen la capacitación y el proyecto de estudio documentados; (2) microtaller de Introducción a la Lógica de Programación ya comprometido con Jonny; (3) conferencia con invitado del sector empresarial, gestionada por el profesor Montoya vía Daniel, con fecha límite del 17 de septiembre. Formalizar con Carlos Osorio Ramírez (CIID) la donación de cupos en diplomados como premio de la Hackatón, y sumar un stand físico por cada área de la UIFCE en el pasillo de la biblioteca. Activar la Red de Aliados Académicos y solicitar estudiantes de corresponsabilidad con suficiente anticipación para que la convocatoria a la Hackatón no se anuncie de un día para otro, como ocurrió en la edición cancelada de 2025-2.",
        espera="Primera edición ejecutada con éxito, con mejor convocatoria a la Hackatón que en 2025-2, cuando se canceló por baja inscripción, y con el presupuesto de patrocinios (cupos CIID, refrigerios y bonos vía VIE) confirmado antes del evento.",
        fundamento="El documento técnico, el estudio de viabilidad y el cronograma preliminar ya fueron formulados en 2026-1; falta su ejecución. El video de empalme aporta el detalle operativo completo (estructura de tres días de la Hackatón, contactos institucionales concretos para presupuesto y patrocinios, y la lección de que la baja convocatoria en la edición cancelada se debió a difusión tardía).",
        entregables="* Evento ejecutado.\n* Hackatón, microtaller y conferencia realizados.\n* Cupos CIID y presupuesto VIE confirmados.\n* Informe de resultados.",
        priority=False,
    ),
    dict(
        id="estrategicos-boletin-digital",
        cat="Proyectos estratégicos — NUEVO",
        idea="Boletín Digital UIFCE",
        hacer="Diseñar la plantilla del boletín, definir el protocolo de recolección de contenido con cada área (AA, GC, DS, Virtualización, CL) y publicar tres números durante el semestre.",
        espera="Tres ediciones publicadas con contenido aportado por al menos tres áreas distintas de la UIFCE en cada número.",
        fundamento="Responde al entregable de \"boletines informativos\" que quedó pendiente en la planeación 2026-1 y a la ausencia de un canal consolidado de información transversal entre áreas identificada en el diagnóstico.",
        entregables="* Plantilla aprobada.\n* Protocolo de recolección de contenido.\n* Tres ediciones publicadas.",
        priority=False,
    ),
    dict(
        id="estrategicos-red-aliados-academicos",
        cat="Proyectos estratégicos — NUEVO",
        idea="Red de Aliados Académicos UIFCE",
        hacer="Construir un listado vivo de profesores y asignaturas aliadas por programa curricular (administración, economía, contaduría) que faciliten la convocatoria presencial a eventos y servicios de la UIFCE.",
        espera="Red activa y utilizada como mecanismo formal de convocatoria para la Semana UIFCE y los microtalleres del semestre.",
        fundamento="Generaliza una necesidad puntual ya identificada en la planeación 2026-1 para la convocatoria de la Semana UIFCE, convirtiéndola en un mecanismo permanente en lugar de reconstruirla cada vez que se organiza un evento.",
        entregables="* Base de datos de aliados.\n* Protocolo de activación por evento.\n* Reporte de asistencia atribuible.",
        priority=False,
    ),
    dict(
        id="innovacion-repositorio-ia",
        cat="Innovación y eficiencia operativa — NUEVO",
        idea="Repositorio de material con IA y automatización de producción de video",
        hacer="Evaluar y adoptar herramientas de inteligencia artificial en dos frentes: (1) un repositorio de material clasificado y buscable (piezas, hacks, videos, documentos de proyectos), y (2) automatización de tareas repetitivas de edición de video (subtitulado, recortes por plataforma, primeros cortes). Articular el repositorio con el soporte documental permanente del área (idea siguiente).",
        espera="Un repositorio de material en operación que reduzca el tiempo de búsqueda frente a la práctica actual en Drive, y al menos un flujo de edición de video con una etapa automatizada funcionando de forma regular hacia el final del semestre.",
        fundamento="El diagnóstico identificó que el conocimiento del área se pierde si no se organiza y que toda la edición de video depende hoy de un único monitor de artes trabajando de forma artesanal. La IA es la palanca que permite sostener el principio de calidad sobre cantidad sin aumentar esa carga operativa. (Ver Portafolio de Proyectos 2026-2, numeral 3.4.)",
        entregables="* Repositorio de material clasificado y en operación.\n* Al menos un flujo de edición de video automatizado.\n* Guía interna de uso de las herramientas de IA adoptadas.",
        priority=False,
    ),
    dict(
        id="documentacion-repositorio-permanente",
        cat="Documentación y memoria institucional — NUEVO",
        idea="Repositorio documental permanente de Estrategias Tecnológicas",
        hacer="Diseñar y poner en marcha un repositorio propio de ET (carpeta estructurada en Drive con convención de nombres y metadatos mínimos: fecha, tipo de pieza, campaña, autor) donde se archive de forma permanente todo el material del área —piezas gráficas, videos, documentos de proyectos, credenciales y evidencias de gestión de redes— más allá de cada semestre, en articulación con el repositorio general que ya lidera Gestión del Conocimiento (GC). Adoptar el mecanismo propuesto en el empalme: que cada máster entrante tome accesos directos al repositorio único en lugar de copiar carpetas completas semestre a semestre, y mantener un asistente de IA con contexto del Drive (ej. Gemini) que permita ubicar cualquier archivo por consulta en lenguaje natural.",
        espera="Un soporte documental único y organizado, disponible para cualquier monitor entrante desde el primer día de empalme, que reduzca el riesgo de pérdidas como la de la cuenta de Instagram y evite que el conocimiento del área se disperse entre semestres.",
        fundamento="El diagnóstico identificó como brecha que \"el conocimiento generado se pierde si no se organiza\" (material disperso en Drive sin estructura accesible), y la contingencia de pérdida de la cuenta de Instagram evidencia de forma directa el costo de no contar con un respaldo documental centralizado de accesos, credenciales y activos del área.",
        entregables="* Estructura de carpetas y convención de nombres definida.\n* Migración del material existente.\n* Protocolo de archivo para credenciales y accesos.",
        priority=False,
    ),
    dict(
        id="documentacion-propuesta-drive-gc",
        cat="Documentación y memoria institucional — NUEVO",
        idea="Propuesta de reorganización del Drive UIFCE ante Gestión del Conocimiento (GC)",
        hacer="Presentar a GC la propuesta formal de repositorio documental permanente para todo el Drive de UIFCE: separar el contenido vigente de cada semestre del contenido permanente por proyecto de continuidad, con convención común de nomenclatura y uso de accesos directos en lugar de copias, tomando como piloto la carpeta ya reorganizada de ET 2026-2. Acordar con GC qué proyectos migran primero al repositorio permanente y quién lidera la migración del material disperso ya existente.",
        espera="Aval de GC sobre el modelo propuesto, creación de la carpeta piloto \"ET — Repositorio Permanente\" dentro del Drive general, y un plan de migración acordado en conjunto para el contenido histórico disperso.",
        fundamento="El diagnóstico identificó que el conocimiento del área se pierde si no se organiza, y la reunión de empalme entre la líder saliente y la entrante llegó de forma independiente a la misma conclusión: dejar de copiar carpetas completas semestre a semestre y usar accesos directos hacia un repositorio único por proyecto. Como el Drive general lo administra GC y no ET, esta idea requiere una propuesta formal y su aval antes de poder ejecutarse.",
        entregables="* Estructura de carpetas propuesta para el Drive interno de ET 2026-2.\n* Propuesta formal presentada a GC.\n* Aval de GC y plan de migración acordado.",
        priority=False,
    ),
    dict(
        id="acompanamiento-interarea",
        cat="Acompañamiento interárea",
        idea="Acompañamiento Interárea",
        hacer="Ejecutar la matriz de acompañamiento a Apoyos Académicos, Gestión del Conocimiento, Desarrollo, Virtualización, Cursos Libres y Coordinación bajo el protocolo de solicitud definido: canal único ante el Máster ET, SLA de 5 días hábiles para piezas puntuales y 4 semanas para campañas institucionales, y criterios de priorización ante solicitudes simultáneas.",
        espera="Cumplimiento del SLA en al menos el 80% de las solicitudes, y necesidades de las seis áreas recogidas formalmente en la reunión de planeación de inicio de semestre.",
        fundamento="La Guía de Empalme UIFCE ya reconoce a ET como responsable de redes sociales y difusión y como colaborador en micrositio, certificados y YouTube; este ítem traduce ese reconocimiento en compromisos operativos concretos.",
        entregables="* Registro compartido de solicitudes.\n* Cumplimiento de SLA.\n* Necesidades de cada área recogidas al inicio de semestre.",
        priority=False,
    ),
    dict(
        id="documentacion-memorias-2026-2",
        cat="Documentación y memoria institucional — NUEVO",
        idea="Memorias UIFCE 2026-2S",
        hacer="Construir un repositorio de memorias del semestre 2026-2 del área —piezas destacadas, eventos, métricas de canales, hitos, decisiones y aprendizajes— recopilando el material de forma continua durante el semestre y no al cierre, con una estructura de carpetas y una convención de nombres consistentes con el repositorio documental permanente. Producir, a partir de ese material, un video final que resuma la gestión del área en el semestre (línea de tiempo de logros, cifras y testimonios del equipo), aplicando la línea gráfica vigente y el flujo de aprobación por el grupo Piezas Redes Sociales.",
        espera="Al cierre del semestre, un repositorio de memorias 2026-2S organizado y navegable, y un video final publicado y enlazado desde ese repositorio, que sirva de cierre de gestión y de insumo directo para el empalme con el equipo entrante.",
        fundamento="El diagnóstico identificó como brecha estructural que el conocimiento del área se pierde si no se organiza, y que cada máster entrante reconstruye contexto copiando carpetas completas de Drive. El empalme 2026-2 se apoyó justamente en un video y su transcripción para transferir el conocimiento tácito de la líder saliente. Un repositorio de memorias y un video de cierre sistematizan esa transferencia y refuerzan la línea del repositorio documental permanente.",
        entregables="* Estructura de carpetas y convención de nombres del repositorio de memorias 2026-2S.\n* Recopilación continua de material a lo largo del semestre (piezas, eventos, métricas, aprendizajes).\n* Guion y edición del video final del área.\n* Video final publicado y enlazado en el repositorio de memorias.",
        priority=False,
    ),
    dict(
        id="cursoslibres-piezas-primer-lanzamiento",
        cat="Acompañamiento a Cursos Libres",
        idea="Elaboración de piezas gráficas de Cursos Libres — Primer lanzamiento",
        hacer="Gestionar y producir las piezas gráficas de difusión del primer lanzamiento de Cursos Libres 2026-2 (seis cursos virtualizados), aplicando la línea gráfica modular vigente de CL y el flujo de aprobación por el grupo Piezas Redes Sociales. El lanzamiento corre de la semana 4 a la semana 8-9 (14 de septiembre a 24 de octubre de 2026). Para cada curso: pieza de inscripción y video corto de lanzamiento del monitor. Coordinar con Cursos Libres con anticipación las fechas, textos y códigos definitivos —es el área con mayor fricción de tiempos según el diagnóstico— y bloquear en el calendario editorial las fechas de entrega antes del inicio de inscripciones de cada curso.",
        espera="Las seis piezas de inscripción y sus videos de lanzamiento aprobados y publicados antes del inicio de inscripciones de cada curso, sin retrasos atribuibles a coordinación tardía con Cursos Libres.",
        fundamento="El diagnóstico y el video de empalme señalan a Cursos Libres como el área con la que ET tiene mayor fricción de tiempos: sus agendamientos se demoran y luego se piden piezas con plazos muy cortos. La planeación 2026-1 dejó el despliegue de la línea gráfica modular de CL en curso. Adelantar y estandarizar la producción del primer lanzamiento reduce ese riesgo recurrente.",
        entregables="* Excel Básico Virtualizado (CLEBV-I) — monitor: María Fernanda Celis · sem 4-8 (14 sep - 17 oct) · 5 sesiones · cupo 10-20 · prerrequisito: ninguno.\n* Excel Intermedio Virtualizado (VEA-20202) — monitor: Joel Santiago Rodríguez Guzmán · sem 4-9 (14 sep - 24 oct) · 6 sesiones · cupo 12-24 · prerrequisito: Excel Básico Virtualizado.\n* Econometría en Python (CLEPY202602) — monitor: Laura Angélica Cárdenas Cely · sem 4-8 (14 sep - 17 oct) · 5 sesiones · cupo 10-20 · prerrequisito: Introducción a la Programación o a Lógica, y estar cursando o haber cursado Econometría I.\n* Introducción a la Programación en Python y R (CLIPPYR292602) — monitor: Diego Alejandro Garnica Mamanché · sem 4-8 (14 sep - 17 oct) · 5 sesiones · cupo 10-20 · prerrequisito: ninguno.\n* Introducción a Power BI (CLIPBI202602) — monitor: Paula Sofía Bocarejo Alberto · sem 4-8 (14 sep - 17 oct) · 5 sesiones · cupo 10-20 · prerrequisito: Excel Intermedio Virtualizado.\n* Siigo Nube (CLSIN2020602) — monitor: Jean Carlos Baquero García · sem 4-8 (14 sep - 17 oct) · 5 sesiones · cupo 10-20 · prerrequisito: Contabilidad de Inversión y Financiación.",
        priority=False,
    ),
]

# ---------- Sheet 1: Planeación ----------
ws1 = wb.active
ws1.title = "Planeación"

ws1["A1"] = "ESTRATEGIAS TECNOLÓGICAS — PLANEACIÓN 2026-2"
ws1["A1"].font = TITLE_FONT
ws1.merge_cells("A1:G1")
ws1["A2"] = ("UIFCE · Facultad de Ciencias Económicas · UNAL — Documento para presentar a Coordinación: "
             "qué se debe hacer, qué se espera y con qué fundamento, para cada línea de trabajo de ET en 2026-2. "
             "Principio rector del semestre: calidad sobre cantidad para posicionar la Unidad (informes 2025-1 a 2026-1, "
             "y video de empalme entre líder saliente y entrante).")
ws1["A2"].font = SUB_FONT
ws1.merge_cells("A2:G2")

headers1 = ["ID", "Categoría", "#", "Idea principal", "Qué se debe hacer", "Qué se espera", "Fundamento"]
row0 = 4
for j, htext in enumerate(headers1, start=1):
    c = ws1.cell(row=row0, column=j, value=htext)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

r = row0 + 1
for i, d in enumerate(IDEAS):
    ws1.cell(row=r, column=1, value=d["id"]).alignment = WRAP_TOP
    ws1.cell(row=r, column=1).font = Font(size=9, color="595959")
    ws1.cell(row=r, column=2, value=d["cat"]).alignment = WRAP_TOP
    ws1.cell(row=r, column=2).font = Font(bold=True, size=9, color="7A1F2B")
    ws1.cell(row=r, column=3, value=i + 1).alignment = Alignment(horizontal="center", vertical="top")
    ws1.cell(row=r, column=4, value=d["idea"]).alignment = WRAP_TOP
    ws1.cell(row=r, column=4).font = Font(bold=True, size=10)
    ws1.cell(row=r, column=5, value=d["hacer"]).alignment = WRAP_TOP
    ws1.cell(row=r, column=5).font = Font(size=10)
    ws1.cell(row=r, column=6, value=d["espera"]).alignment = WRAP_TOP
    ws1.cell(row=r, column=6).font = Font(size=10)
    ws1.cell(row=r, column=7, value=d["fundamento"]).alignment = WRAP_TOP
    ws1.cell(row=r, column=7).font = Font(size=10, italic=True, color="595959")
    for col in range(1, 8):
        cell = ws1.cell(row=r, column=col)
        cell.border = BORDER
        if d.get("priority"):
            cell.fill = PRIORITY_FILL
        elif i % 2 == 1:
            cell.fill = ALT_FILL
    r += 1

ws1.column_dimensions["A"].width = 24
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 5
ws1.column_dimensions["D"].width = 26
ws1.column_dimensions["E"].width = 50
ws1.column_dimensions["F"].width = 40
ws1.column_dimensions["G"].width = 45
ws1.freeze_panes = "A5"
for rr in range(row0 + 1, r):
    ws1.row_dimensions[rr].height = 130

# ---------- Sheet 2: Planeación del Área ----------
ws2 = wb.create_sheet("Planeación del Área")
ws2["A1"] = "PLANEACIÓN DEL ÁREA — DETALLE OPERATIVO 2026-2"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:G1")
ws2["A2"] = ("Mismo contenido de la hoja \"Planeación\", organizado por actividad con entregables explícitos, enriquecido con "
             "los hallazgos operativos concretos (contactos, procesos y pendientes) del video de empalme entre la líder saliente "
             "y la entrante — ver Transcripción y Notas del Empalme ET 2026-2, numeral 15. "
             "Principio rector del semestre: calidad sobre cantidad para posicionar la Unidad. "
             "Completar en equipo con fechas y responsables de pago al inicio del semestre.")
ws2["A2"].font = SUB_FONT
ws2.merge_cells("A2:G2")

headers2 = ["ID", "Categoría", "Actividad", "Qué se debe hacer", "Qué se espera", "Fundamento", "Entregables"]
row0b = 4
for j, htext in enumerate(headers2, start=1):
    c = ws2.cell(row=row0b, column=j, value=htext)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

r2 = row0b + 1
for i, d in enumerate(IDEAS):
    ws2.cell(row=r2, column=1, value=d["id"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=1).font = Font(size=9, color="595959")
    ws2.cell(row=r2, column=2, value=d["cat"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=2).font = Font(bold=True, size=9, color="7A1F2B")
    ws2.cell(row=r2, column=3, value=d["idea"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=3).font = Font(bold=True, size=10)
    ws2.cell(row=r2, column=4, value=d["hacer"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=4).font = Font(size=10)
    ws2.cell(row=r2, column=5, value=d["espera"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=5).font = Font(size=10)
    ws2.cell(row=r2, column=6, value=d["fundamento"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=6).font = Font(size=10, italic=True, color="595959")
    ws2.cell(row=r2, column=7, value=d["entregables"]).alignment = WRAP_TOP
    ws2.cell(row=r2, column=7).font = Font(size=10)
    for col in range(1, 8):
        cell = ws2.cell(row=r2, column=col)
        cell.border = BORDER
        if d.get("priority"):
            cell.fill = PRIORITY_FILL
        elif i % 2 == 1:
            cell.fill = ALT_FILL
    r2 += 1

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 26
ws2.column_dimensions["D"].width = 48
ws2.column_dimensions["E"].width = 38
ws2.column_dimensions["F"].width = 42
ws2.column_dimensions["G"].width = 32
ws2.freeze_panes = "A5"
for rr in range(row0b + 1, r2):
    ws2.row_dimensions[rr].height = 130

# ---------- Notas de leyenda ----------
for ws, lastrow in ((ws1, r), (ws2, r2)):
    note_row = lastrow + 1
    ws.cell(row=note_row, column=1, value="Relleno naranja = línea prioritaria/crítica del semestre (Instagram y LinkedIn).").font = Font(italic=True, size=9, color="7A1F2B")

for ws in (ws1, ws2):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "4:4"

wb.save("/tmp/outputs/Estrategias Tecnológicas-Planeación 2026-2.xlsx")
print("OK planeacion 2026-2 v2 — filas:", len(IDEAS))
